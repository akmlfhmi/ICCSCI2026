import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
from datetime import datetime
import statistics
import matplotlib
matplotlib.use('TkAgg')

# Try to import optional dependencies
try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from scapy.all import rdpcap
    SCAPY_AVAILABLE = True
except ImportError:
    SCAPY_AVAILABLE = False

try:
    from fpdf import FPDF

    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False

try:
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

# 1. Add a helper function for average calculation near the top (after imports, before class):
def calculate_average(values):
    present = [float(v) for v in values if v not in ('', None, '0.00', 0.0)]
    if not present:
        return "0.00"
    return f"{sum(present)/len(present):.2f}"


class NetworkPerformanceExtractor:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Network Performance Extraction Tool")
        self.root.geometry("800x600")
        self.root.configure(bg='#f0f0f0')

        # Check dependencies
        self.check_dependencies()

        # Variables to store file paths (only file name shown)
        self.novpn_files = [tk.StringVar() for _ in range(5)]
        self.s2s_files = [tk.StringVar() for _ in range(5)]
        self.dmvpn_files = [tk.StringVar() for _ in range(5)]
        self.s2sdmvpn_files = [tk.StringVar() for _ in range(5)]

        # Store full file paths for processing
        self.file_full_paths = {  # key: (topology, idx) -> full path
            **{('NoVPN', i): '' for i in range(5)},
            **{('S2S VPN', i): '' for i in range(5)},
            **{('DMVPN', i): '' for i in range(5)},
            **{('S2S&DMVPN', i): '' for i in range(5)}
        }

        # Results storage
        self.results = {}

        # Create main container
        self.main_container = tk.Frame(self.root, bg='#f0f0f0')
        self.main_container.pack(expand=True, fill='both')

        # Setup initial GUI
        self.setup_file_selection_gui()

    def check_dependencies(self):
        """Check if required dependencies are installed"""
        missing_deps = []

        if not SCAPY_AVAILABLE:
            missing_deps.append("scapy")
        if not PANDAS_AVAILABLE:
            missing_deps.append("pandas")
        if not FPDF_AVAILABLE:
            missing_deps.append("fpdf2")

        if missing_deps:
            deps_str = ", ".join(missing_deps)
            messagebox.showwarning(
                "Missing Dependencies",
                f"The following packages are not installed: {deps_str}\n\n"
                f"Please install them using:\n"
                f"pip install {' '.join(missing_deps)}\n\n"
                f"Some features may not work properly."
            )

    def clear_container(self):
        """Clear all widgets from the main container"""
        for widget in self.main_container.winfo_children():
            widget.destroy()

    def setup_file_selection_gui(self):
        """Setup the file selection interface"""
        self.clear_container()

        # Title
        title_label = tk.Label(self.main_container, text="Network Performance Extraction Tool",
                               font=("Arial", 16, "bold"), bg='#f0f0f0')
        title_label.pack(pady=(8, 8))

        # Instruction label (replace subtitle)
        instruction_label = tk.Label(self.main_container,
                                     text="Please select .pcap or .pcapng files to be extracted",
                                     font=("Arial", 10), bg='#f0f0f0', fg='#666')
        instruction_label.pack(pady=(0, 5))

        # Main frame
        main_frame = tk.Frame(self.main_container, bg='#f0f0f0')
        main_frame.pack(expand=True, fill='both', padx=20, pady=(0, 10))

        # File selection section as a 2x2 grid
        file_frame = tk.LabelFrame(main_frame, text="Select Files",
                                   font=("Arial", 12, "bold"), bg='#f0f0f0', padx=10, pady=10)
        file_frame.pack(pady=10)

        topologies = [
            ('NoVPN', 'Topology 1: No VPN', self.novpn_files),
            ('S2S VPN', 'Topology 2: S2S VPN', self.s2s_files),
            ('DMVPN', 'Topology 3: DMVPN', self.dmvpn_files),
            ('S2S&DMVPN', 'Topology 4: S2S VPN & DMVPN', self.s2sdmvpn_files)
        ]
        # Arrange in 2x2 grid
        for idx, (topo_key, topo_label, file_vars) in enumerate(topologies):
            grid_row = idx // 2
            grid_col = idx % 2
            block = tk.Frame(file_frame, bg='#f0f0f0', padx=20, pady=10, relief='groove', bd=1)
            block.grid(row=grid_row, column=grid_col, padx=30, pady=15, sticky='n')
            label = tk.Label(block, text=topo_label, font=("Arial", 11, "bold"), bg='#f0f0f0')
            label.pack(pady=(0, 8))
            for row in range(5):
                row_frame = tk.Frame(block, bg='#f0f0f0')
                row_frame.pack(fill='x', pady=2)
                row_label = tk.Label(row_frame, text=f"File {row+1}:", font=("Arial", 10), bg='#f0f0f0', anchor='e', width=7)
                row_label.pack(side='left')
                entry = tk.Entry(row_frame, textvariable=file_vars[row], font=("Arial", 10), width=28, state='readonly')
                entry.pack(side='left', padx=(10, 2))
                btn = tk.Button(row_frame, text="Select File", command=lambda v=file_vars[row], k=(topo_key, row): self.select_file(v, k),
                                font=("Arial", 10), bg='#2196F3', fg='white', padx=8)
                btn.pack(side='left', padx=(2, 0))

        # Buttons frame
        buttons_frame = tk.Frame(main_frame, bg='#f0f0f0')
        buttons_frame.pack(pady=20)

        # Extract button
        extract_btn = tk.Button(buttons_frame, text="Extract Network Performance",
                                command=self.extract_performance,
                                font=("Arial", 12, "bold"),
                                bg='#4CAF50', fg='white', padx=20, pady=10)
        extract_btn.pack(side='left', padx=10)

        # Close button for first page
        close_btn = tk.Button(buttons_frame, text="✕ Close", command=self.close_application,
                              font=("Arial", 10, "bold"), bg='#F44336', fg='white', padx=20, pady=10)
        close_btn.pack(side='left', padx=10)

        # Progress bar
        self.progress = ttk.Progressbar(main_frame, length=400, mode='determinate')
        self.progress.pack(pady=10)

        # Status label
        self.status_label = tk.Label(main_frame, text="Ready to extract network performance metrics",
                                     font=("Arial", 10), bg='#f0f0f0', fg='#666')
        self.status_label.pack(pady=5)

    def create_file_selector(self, parent, label_text, var, key):
        frame = tk.Frame(parent, bg='#f0f0f0')
        frame.pack(fill='x', pady=5)

        label = tk.Label(frame, text=label_text, font=("Arial", 10), bg='#f0f0f0', width=20, anchor='w')
        label.pack(side='left')

        entry = tk.Entry(frame, textvariable=var, font=("Arial", 10), width=50, state='readonly')
        entry.pack(side='left', padx=10, expand=True, fill='x')

        btn = tk.Button(frame, text="Select Files", command=lambda: self.select_file(var, key),
                        font=("Arial", 10), bg='#2196F3', fg='white', padx=10)
        btn.pack(side='right')

    def select_file(self, var, key):
        filename = filedialog.askopenfilename(
            title="Select PCAP file",
            filetypes=[("All files", "*.*"), ("PCAP files", "*.pcap"), ("PCAPNG files", "*.pcapng")]
        )
        if filename:
            var.set(os.path.basename(filename))  # Show only the file name
            self.file_full_paths[key] = filename  # Store the full path for processing

    def extract_performance(self):
        topologies = ['NoVPN', 'S2S VPN', 'DMVPN', 'S2S&DMVPN']
        selected_files = {topo: [] for topo in topologies}
        for topo in topologies:
            for idx in range(5):
                file_path = self.file_full_paths.get((topo, idx), '')
                if file_path:
                    selected_files[topo].append((idx, file_path))
        if not any(selected_files[topo] for topo in topologies):
            messagebox.showerror("Error", "Please select at least one PCAP file.")
            return
        try:
            self.progress['value'] = 0
            self.status_label.config(text="Processing PCAP files...")
            self.root.update()
            total_files = sum(len(files) for files in selected_files.values())
            processed = 0
            self.results = {topo: [None, None, None, None, None] for topo in topologies}
            for topo in topologies:
                for idx, file_path in selected_files[topo]:
                    self.status_label.config(text=f"Analyzing {topo} File {idx+1}...")
                    self.root.update()
                    metrics = self.analyze_pcap(file_path)
                    self.results[topo][idx] = metrics
                    processed += 1
                    self.progress['value'] = (processed / total_files) * 100
                    self.root.update()
            self.progress['value'] = 100
            self.status_label.config(text="Analysis complete!")
            print("DEBUG after extraction:", self.results)  # Debug print
            self.show_results()
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            self.status_label.config(text="Error occurred during analysis")

    def analyze_pcap(self, file_path):
        """Analyze PCAP file and extract network performance metrics"""
        if not SCAPY_AVAILABLE:
            raise Exception("Scapy library is not installed. Please install it using: pip install scapy")

        try:
            import scapy.layers.inet
            packets = rdpcap(file_path)

            # Initialize metrics (allow float values)
            metrics = {
                'Throughput (kbps)': 0.0,
                'Latency (ms)': 0.0,
                'Bandwidth (Mbps)': 0.0,
                'Jitter (ms)': 0.0,
                'Packet Loss (%)': 0.0
            }

            if not packets:
                return metrics

            # Extract packet data
            packet_times = []
            packet_sizes = []
            tcp_packets = []
            udp_packets = []

            for packet in packets:
                packet_times.append(packet.time)
                packet_sizes.append(len(packet))

                if packet.haslayer(scapy.layers.inet.TCP):
                    tcp_packets.append(packet)
                elif packet.haslayer(scapy.layers.inet.UDP):
                    udp_packets.append(packet)

            if len(packet_times) < 2:
                return metrics

            # Calculate duration
            duration = packet_times[-1] - packet_times[0]
            if duration <= 0:
                duration = 1

            # Calculate Throughput (kbps)
            total_bytes = sum(packet_sizes)
            throughput_bps = (total_bytes * 8) / duration  # bits per second
            metrics['Throughput (kbps)'] = throughput_bps / 1000  # Convert to kbps

            # Calculate Bandwidth (Mbps)
            metrics['Bandwidth (Mbps)'] = throughput_bps / 1_000_000  # Convert to Mbps

            # Calculate Latency (ms) - Average inter-packet delay
            if len(packet_times) > 1:
                delays = [packet_times[i] - packet_times[i - 1] for i in range(1, len(packet_times))]
                avg_delay = statistics.mean(delays) * 1000  # Convert to ms
                metrics['Latency (ms)'] = avg_delay

                # Calculate Jitter (ms) - Standard deviation of delays
                if len(delays) > 1:
                    jitter = statistics.stdev(delays) * 1000  # Convert to ms
                    metrics['Jitter (ms)'] = jitter

            # Calculate Packet Loss (%) - Simple estimation based on sequence gaps
            packet_loss_percentage = 0
            if tcp_packets:
                # For TCP packets, check for retransmissions or sequence gaps
                seq_numbers = []
                for tcp_pkt in tcp_packets:
                    if tcp_pkt.haslayer(scapy.layers.inet.TCP):
                        seq_numbers.append(tcp_pkt[scapy.layers.inet.TCP].seq)

                if seq_numbers:
                    expected_packets = len(set(seq_numbers))
                    actual_packets = len(seq_numbers)
                    if expected_packets > 0:
                        packet_loss_percentage = max(0, ((actual_packets - expected_packets) / actual_packets) * 100)

            metrics['Packet Loss (%)'] = packet_loss_percentage

            # Round all values to 2 decimal places
            for key in metrics:
                metrics[key] = round(metrics[key], 2)

            return metrics

        except Exception as e:
            raise Exception(f"Error analyzing PCAP file {file_path}: {str(e)}")

    def show_results(self):
        print("DEBUG in show_results:", self.results)  # Debug print
        self.clear_container()
        main_frame = tk.Frame(self.main_container, bg='#f0f0f0')
        main_frame.pack(expand=True, fill='both', padx=20, pady=20)
        title_label = tk.Label(main_frame, text="Network Performance Analysis Results",
                               font=("Arial", 18, "bold"), bg='#f0f0f0')
        title_label.pack(pady=(0, 10))
        # Table frame
        table_frame = tk.Frame(main_frame, bg='#f0f0f0')
        table_frame.pack(expand=True, fill='both')

        columns = ('Topology', 'Value', 'Throughput (kbps)', 'Latency (ms)', 'Bandwidth (Mbps)', 'Jitter (ms)', 'Packet Loss (%)')
        tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=30)
        for col in columns:
            tree.heading(col, text=col)
        tree.column('Topology', width=120, anchor='center')
        tree.column('Value', width=70, anchor='center')
        tree.column('Throughput (kbps)', width=120, anchor='center')
        tree.column('Latency (ms)', width=120, anchor='center')
        tree.column('Bandwidth (Mbps)', width=120, anchor='center')
        tree.column('Jitter (ms)', width=120, anchor='center')
        tree.column('Packet Loss (%)', width=120, anchor='center')
        style = ttk.Style()
        style.configure('Treeview', rowheight=28)
        style.configure('Treeview.Heading', font=("Arial", 10, "bold"))
        style.layout('Treeview', [('Treeview.treearea', {'sticky': 'nswe'})])
        style.configure('Treeview', borderwidth=2, relief='solid')
        style.map('Treeview', background=[('selected', '#e0e0e0')])
        style.configure('Treeview', highlightthickness=2, bd=2, font=("Arial", 10))
        style.configure('Treeview', fieldbackground='#f0f0f0')
        style.configure('Treeview', bordercolor='#222')
        style.configure('Treeview', borderwidth=2)
        style.configure('Treeview', relief='solid')
        # Custom tag for bold average row
        tree.tag_configure('average_bold', font=("Arial", 10, "bold"))

        topologies = ['NoVPN', 'S2S VPN', 'DMVPN', 'S2S&DMVPN']
        topology_names_gui = {
            'NoVPN': 'No VPN',
            'S2S VPN': 'S2S VPN',
            'DMVPN': 'DMVPN',
            'S2S&DMVPN': 'S2S VPN & DMVPN'
        }
        metrics_list = [
            'Throughput (kbps)',
            'Latency (ms)',
            'Bandwidth (Mbps)',
            'Jitter (ms)',
            'Packet Loss (%)'
        ]
        has_data = False
        row_indices = []  # For drawing lines
        for topo in topologies:
            display_name = topology_names_gui.get(topo, topo)
            results_list = self.results.get(topo, [None]*5)
            # Prepare per-metric lists for each value
            metric_rows = []
            for idx in range(5):
                row = []
                result = results_list[idx]
                for metric in metrics_list:
                    val = result.get(metric, '') if result else ''
                    # Format to 2 decimal places, including 0.00
                    try:
                        val = float(val)
                        val = f"{val:.2f}"
                    except Exception:
                        val = "0.00" if val in ('', None) else str(val)
                    row.append(val)
                metric_rows.append(row)
            # Check if any value is not empty
            if any(any(v not in ('', None, '0.00') for v in row) for row in metric_rows):
                has_data = True
            # Calculate averages using only present values
            avg_row = []
            for m_idx, metric in enumerate(metrics_list):
                vals = [float(metric_rows[i][m_idx]) for i in range(5) if metric_rows[i][m_idx] not in ('', None, '0.00')]
                avg = calculate_average(vals)
                avg_row.append(avg)
            # Insert rows: 5 values, then average
            for idx in range(5):
                # Topology name only in the 3rd row (index 2), but not bolded
                topo_cell = display_name if idx == 2 else ''
                tree.insert('', 'end', values=(topo_cell, str(idx+1),
                                               metric_rows[idx][0], metric_rows[idx][1], metric_rows[idx][2], metric_rows[idx][3], metric_rows[idx][4]))
            # Insert average row, bold all columns
            tree.insert('', 'end', values=(' ', 'Average', avg_row[0], avg_row[1], avg_row[2], avg_row[3], avg_row[4]), tags=('average_bold',))
            # Record the index for drawing a line after this group
            row_indices.append(tree.get_children()[-1])
        if not has_data:
            tree.insert('', 'end', values=("No data available", '', '', '', '', '', ''))

        # Add vertical scrollbar
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side='right', fill='y')
        tree.pack(expand=True, fill='both', side='left')

        # Draw 1px gray line after each group
        def draw_group_lines():
            for item in table_frame.place_slaves():
                if isinstance(item, tk.Canvas):
                    item.destroy()
            tree.update_idletasks()
            children = tree.get_children()
            for row_id in row_indices:
                bbox = tree.bbox(row_id)
                if bbox:
                    y = bbox[1] + bbox[3]
                    line_canvas = tk.Canvas(table_frame, width=tree.winfo_width(), height=1, bg='#f0f0f0', highlightthickness=0, bd=0)
                    line_canvas.create_line(0, 0, tree.winfo_width(), 0, fill='#cccccc', width=1)
                    line_canvas.place(x=0, y=y)
        tree.bind('<Expose>', lambda e: draw_group_lines())
        tree.bind('<Configure>', lambda e: draw_group_lines())
        tree.after(100, draw_group_lines)

        # Navigation and export buttons as before
        btn_frame = tk.Frame(self.main_container, bg='#f0f0f0')
        btn_frame.pack(fill='x', padx=20, pady=10, side='bottom')
        left_btn_frame = tk.Frame(btn_frame, bg='#f0f0f0')
        left_btn_frame.pack(side='left')
        select_new_btn = tk.Button(left_btn_frame, text="Select New Files", command=self.setup_file_selection_gui,
                                   font=("Arial", 7, "bold"), bg='#2196F3', fg='white', padx=14, pady=5)
        select_new_btn.pack(side='left', padx=2)
        back_btn = tk.Button(left_btn_frame, text="← Back", command=self.setup_file_selection_gui,
                             font=("Arial", 7, "bold"), bg='#9E9E9E', fg='white', padx=14, pady=5)
        back_btn.pack(side='left', padx=2)
        close_btn = tk.Button(left_btn_frame, text="✕ Close", command=self.close_application,
                              font=("Arial", 7, "bold"), bg='#F44336', fg='white', padx=14, pady=5)
        close_btn.pack(side='left', padx=2)
        export_frame = tk.Frame(btn_frame, bg='#f0f0f0')
        export_frame.pack(side='right')
        pdf_btn = tk.Button(export_frame, text="Save Results to PDF", command=self.save_results_to_pdf,
                            font=("Arial", 8), bg='#FF5722', fg='white', padx=14, pady=6)
        pdf_btn.pack(side='left', padx=5)
        excel_btn = tk.Button(export_frame, text="Save Results to Excel", command=self.save_results_to_excel,
                              font=("Arial", 8), bg='#9C27B0', fg='white', padx=14, pady=6)
        excel_btn.pack(side='left', padx=5)
        graph_btn = tk.Button(export_frame, text="Generate Graph", command=self.show_graphs,
                              font=("Arial", 8), bg='#4CAF50', fg='white', padx=14, pady=6)
        graph_btn.pack(side='left', padx=5)

    def close_application(self):
        """Close the application"""
        self.root.quit()
        self.root.destroy()

    def generate_graph_figures(self):
        import matplotlib.pyplot as plt
        import pprint
        print('DEBUG self.results:')
        pprint.pprint({topo: self.results.get(topo, [None]*5) for topo in ['DMVPN', 'S2S VPN', 'NoVPN', 'S2S&DMVPN']})
        metrics_list = [
            ('Throughput (kbps)', 'Throughput (kbps)'),
            ('Latency (ms)', 'Latency (ms)'),
            ('Bandwidth (Mbps)', 'Bandwidth (Mbps)'),
            ('Jitter (ms)', 'Jitter (ms)'),
            ('Packet Loss (%)', 'Packet Loss (%)')
        ]
        # Order: Throughput, Bandwidth, Jitter, Packet Loss, Latency
        topologies = ['DMVPN', 'S2S VPN', 'NoVPN', 'S2S&DMVPN']
        colors = ['#e41a1c', '#377eb8', '#4daf4a', '#ff9800']
        display_names = ['DMVPN', 'S2S VPN', 'No VPN', 'S2S VPN & DMVPN']
        self.graph_figures = []
        for i, (metric, label) in enumerate(metrics_list):
            # Debug print for Packet Loss values for each topology
            if metric == 'Packet Loss (%)':
                print('DEBUG Packet Loss Results:', {topo: [r.get('Packet Loss (%)') if r else None for r in self.results.get(topo, [None]*5)] for topo in topologies})
            fig, ax = plt.subplots(figsize=(7, 4))
            x_pos = list(range(1, 6))
            all_y_vals = []
            all_x_vals = []
            # For packet loss, collect all y for all topologies to check if all are the same
            packet_loss_all_y = []
            for t_idx, topo in enumerate(topologies):
                results_list = self.results.get(topo, [None]*5)
                y_vals = []
                x_vals = []
                for file_idx in range(5):
                    result = results_list[file_idx]
                    if result is not None:
                        val = result.get(metric, None)
                        if val is not None and val != '':
                            y_vals.append(float(val))
                            # For Bandwidth, add a small offset to x to avoid overlap with Latency
                            if metric == 'Bandwidth (Mbps)':
                                x_vals.append(file_idx+1 + 0.12)
                            else:
                                x_vals.append(file_idx+1)
                if metric == 'Packet Loss (%)':
                    packet_loss_all_y.append(y_vals)
                else:
                    if y_vals:
                        ax.plot(x_vals, y_vals, marker='o', color=colors[t_idx], label=display_names[t_idx], linewidth=2)
                all_y_vals.append(y_vals)
                all_x_vals.append(x_vals)
            # Special handling for Packet Loss: always show all topologies in the legend and plot only real data
            if metric == 'Packet Loss (%)':
                # Always add a dummy plot for each topology to force legend entries
                for t_idx, topo in enumerate(topologies):
                    ax.plot([], [], marker='o', color=colors[t_idx], label=display_names[t_idx], linewidth=2)
                for t_idx, topo in enumerate(topologies):
                    results_list = self.results.get(topo, [None]*5)
                    x_vals = []
                    y_vals = []
                    for file_idx in range(5):
                        result = results_list[file_idx]
                        if result is not None and metric in result:
                            val = result.get(metric, 0)
                            if val == '' or val is None:
                                val = 0
                            x_vals.append(file_idx+1)
                            y_vals.append(float(val))
                    if x_vals:  # Only plot if there is at least one data point
                        ax.plot(x_vals, y_vals, marker='o', color=colors[t_idx], linewidth=2)
            ax.set_xticks(x_pos)
            ax.set_xticklabels([str(i) for i in x_pos])
            ax.set_xlabel('Test')
            ax.set_ylabel(label)
            ax.set_title(label)
            ax.spines['left'].set_visible(True)
            ax.spines['left'].set_color('black')
            ax.spines['bottom'].set_visible(True)
            ax.spines['bottom'].set_color('black')
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            all_vals = []
            for t in topologies:
                results_list = self.results.get(t, [None]*5)
                for i in range(5):
                    result = results_list[i]
                    if result is not None:
                        v = result.get(metric, None)
                        if v is not None and v != '':
                            all_vals.append(float(v))
            if all_vals:
                min_val = min(float(v) for v in all_vals)
                max_val = max(float(v) for v in all_vals)
                margin = (max_val - min_val) * 0.2 if min_val != max_val else 1
                if metric == 'Packet Loss (%)' or metric == 'Bandwidth (Mbps)':
                    ax.set_ylim(0, max_val + margin)
                elif min_val == max_val:
                    ax.set_ylim(min_val - 1, max_val + 1)
                else:
                    ax.set_ylim(min(0, min_val - margin), max_val + margin)
            ax.grid(True, axis='y', linestyle='--', alpha=0.5)
            # Move legend to a position centered between previous and current
            ax.legend(loc='lower right', bbox_to_anchor=(1.075, -0.45), fontsize=11, frameon=True, ncol=2)
            self.graph_figures.append(fig)
            fig.subplots_adjust(bottom=0.25, top=0.82)

    def save_results_to_pdf(self):
        if not FPDF_AVAILABLE:
            messagebox.showerror("Error", "FPDF library is not installed. Please install it using: pip install fpdf2")
            return
        if not MATPLOTLIB_AVAILABLE:
            messagebox.showerror("Error", "matplotlib is not installed. Please install it using: pip install matplotlib")
            return
        # Ensure graphs are generated
        if not hasattr(self, 'graph_figures') or not self.graph_figures:
            self.generate_graph_figures()
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                title="Save results as PDF"
            )
            if not filename:
                return
            import tempfile
            import os
            pdf = FPDF()
            # --- Page 1: Table(s) ---
            pdf.add_page()
            pdf.set_font("Arial", "B", 16)
            pdf.cell(0, 10, "Network Performance Analysis Result", ln=True, align='C')
            pdf.ln(5)
            topology_names_pdf = {
                'NoVPN': 'No VPN',
                'S2S VPN': 'S2S VPN',
                'DMVPN': 'DMVPN',
                'S2S&DMVPN': 'S2S VPN &\nDMVPN'
            }
            # For GUI and Excel, always use single-line name for topology 4
            topology_names_excel = {
                'NoVPN': 'No VPN',
                'S2S VPN': 'S2S VPN',
                'DMVPN': 'DMVPN',
                'S2S&DMVPN': 'S2S VPN & DMVPN'
            }
            metrics_list = [
                'Throughput (kbps)',
                'Latency (ms)',
                'Bandwidth (Mbps)',
                'Jitter (ms)',
                'Packet loss (%)'
            ]
            # Define topologies for PDF export
            topologies = ['NoVPN', 'S2S VPN', 'DMVPN', 'S2S&DMVPN']
            # Table header (single-line cells, all columns visible)
            col_widths = [28, 14, 32, 32, 32, 32, 32]
            table_width = sum(col_widths)
            page_width = pdf.w - pdf.l_margin - pdf.r_margin
            left_margin = max(5, (page_width - table_width) // 2)
            pdf.set_left_margin(left_margin)
            pdf.set_x(left_margin)
            pdf.set_font("Arial", "B", 11)
            headers = [
                "Topology", "Value", "Throughput/kbps", "Latency/ms", "Bandwidth/Mbps", "Jitter/ms", "Packet loss/%"
            ]
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 8, header, border=1, align='C')
            pdf.ln(8)
            pdf.set_font("Arial", "", 10)
            # Print all rows for all topologies in one table, visually merge topology column
            for topo in topologies:
                display_name = topology_names_pdf.get(topo, topo)
                results_list = self.results.get(topo, [None]*5)
                metric_rows = []
                for idx in range(5):
                    row = []
                    result = results_list[idx]
                    for metric in metrics_list:
                        val = result.get(metric, '') if result else ''
                        try:
                            val = float(val)
                            val = f"{val:.2f}"
                        except Exception:
                            val = "0.00" if val in ('', None) else str(val)
                        row.append(val)
                    metric_rows.append(row)
                # Calculate averages using only present values
                avg_row = []
                for m_idx, metric in enumerate(metrics_list):
                    vals = [metric_rows[i][m_idx] for i in range(5) if metric_rows[i][m_idx] not in ('', None, '0.00')]
                    avg = calculate_average(vals)
                    avg_row.append(avg)
                # Print merged topology cell (multi_cell for S2S&DMVPN, single cell for others)
                y_start = pdf.get_y()
                pdf.set_x(left_margin)
                pdf.set_font("Arial", "", 10)
                if topo == 'S2S&DMVPN':
                    # Center the two-line label vertically in the merged cell
                    cell_height = 8 * 6
                    text_lines = ['S2S VPN &', 'DMVPN']
                    text_height = 8 * len(text_lines)
                    y_offset = y_start + (cell_height - text_height) // 2
                    pdf.set_y(y_offset)
                    x0 = left_margin
                    y0 = y_start
                    pdf.multi_cell(col_widths[0], 8, 'S2S VPN &\nDMVPN', border=0, align='C')
                    x1 = x0 + col_widths[0]
                    # Draw all borders manually for a solid box
                    pdf.line(x0, y0, x0, y0 + cell_height)  # left border
                    pdf.line(x1, y0, x1, y0 + cell_height)  # right border
                    pdf.line(x0, y0, x1, y0)  # top border
                    pdf.line(x0, y0 + cell_height, x1, y0 + cell_height)  # bottom border
                    y_topo = y_start
                else:
                    pdf.cell(col_widths[0], 8*6, display_name, border=1, align='C', ln=0)
                    y_topo = y_start
                for idx in range(6):
                    pdf.set_xy(left_margin + col_widths[0], y_topo + idx*8)
                    if idx < 5:
                        pdf.cell(col_widths[1], 8, str(idx+1), border=1, align='C')
                        for m in range(5):
                            pdf.cell(col_widths[m+2], 8, metric_rows[idx][m], border=1, align='C')
                        pdf.ln(0)
                    else:
                        pdf.set_font("Arial", "B", 10)
                        pdf.cell(col_widths[1], 8, 'Average', border=1, align='C')
                        for m in range(5):
                            pdf.cell(col_widths[m+2], 8, avg_row[m], border=1, align='C')
                        pdf.set_font("Arial", "", 10)
                        pdf.ln(0)
                # Draw a horizontal line to divide each topology group
                pdf.set_xy(left_margin, y_topo + 8*6)
                pdf.cell(table_width, 0, '', border='T')
                pdf.set_y(y_topo + 8*6)

            # --- Page 2: All 5 graphs ---
            pdf.add_page()
            pdf.set_font("Arial", "B", 14)
            pdf.cell(0, 10, "Network Performance Graph Analysis", ln=True, align='C')
            pdf.ln(2)
            temp_files = []
            graphs_per_row = 2
            graph_width = 90
            graph_height = 60
            # Center the graphs horizontally
            page_width = pdf.w - pdf.l_margin - pdf.r_margin
            total_graphs_width = graphs_per_row * graph_width + (graphs_per_row - 1) * 20
            left_x = max(5, (page_width - total_graphs_width) // 2 + pdf.l_margin - 10)  # shift 10 units more left
            # Place first 4 graphs in 2x2 grid, last (Jitter) centered below but left-aligned
            for i, fig in enumerate(self.graph_figures):
                with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmpfile:
                    fig.savefig(tmpfile.name, bbox_inches='tight', dpi=100)
                    temp_files.append(tmpfile.name)
                    if i < 4:
                        x = left_x + (i % graphs_per_row) * (graph_width + 20)
                        y = 30 + (i // graphs_per_row) * (graph_height + 30)
                        pdf.image(tmpfile.name, x=x, y=y, w=graph_width, h=graph_height)
                    else:
                        # Place the last graph (Jitter) left-aligned below
                        x = left_x
                        y = 30 + 2 * (graph_height + 30)
                        pdf.image(tmpfile.name, x=x, y=y, w=graph_width, h=graph_height)
            pdf.output(filename)
            for tmp in temp_files:
                try:
                    os.remove(tmp)
                except Exception:
                    pass
            messagebox.showinfo("Success", f"Results saved to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save results to PDF: {str(e)}")

    def save_results_to_excel(self):
        if not PANDAS_AVAILABLE:
            messagebox.showerror("Error", "Pandas library is not installed. Please install it using: pip install pandas openpyxl")
            return
        if not MATPLOTLIB_AVAILABLE:
            messagebox.showerror("Error", "matplotlib is not installed. Please install it using: pip install matplotlib")
            return
        # Ensure graphs are generated
        if not hasattr(self, 'graph_figures') or not self.graph_figures:
            self.generate_graph_figures()
        try:
            from openpyxl.utils import get_column_letter
            def autofit_columns(ws):
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # Get the column name
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except Exception:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column].width = adjusted_width
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save results as Excel"
            )
            if not filename:
                return
            import openpyxl
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            import tempfile
            import os
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Network Performance Results'
            # Table header
            row = 1
            headers = ['Topology', 'Value', 'Throughput (kbps)', 'Latency (ms)', 'Bandwidth (Mbps)', 'Jitter (ms)', 'Packet Loss (%)']
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header).font = Font(bold=True)
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = thin_border
            row += 1
            # Table rows
            # For GUI and Excel, always use single-line name for topology 4
            topology_names_excel = {
                'NoVPN': 'No VPN',
                'S2S VPN': 'S2S VPN',
                'DMVPN': 'DMVPN',
                'S2S&DMVPN': 'S2S VPN & DMVPN'
            }
            metrics_list = [
                'Throughput (kbps)',
                'Latency (ms)',
                'Bandwidth (Mbps)',
                'Jitter (ms)',
                'Packet Loss (%)'
            ]
            topologies = ['NoVPN', 'S2S VPN', 'DMVPN', 'S2S&DMVPN']
            light_gray_fill = PatternFill(start_color='FFF5F5F5', end_color='FFF5F5F5', fill_type='solid')
            for topo in topologies:
                display_name = topology_names_excel.get(topo, topo)
                results_list = self.results.get(topo, [None]*5)
                metric_rows = []
                for idx in range(5):
                    row_vals = []
                    result = results_list[idx]
                    for metric in metrics_list:
                        val = result.get(metric, '') if result else ''
                        try:
                            val = float(val)
                        except Exception:
                            val = 0.0
                        row_vals.append(val)
                    metric_rows.append(row_vals)
                # Calculate averages using only present values
                avg_row = []
                for m_idx, metric in enumerate(metrics_list):
                    vals = [metric_rows[i][m_idx] for i in range(5) if metric_rows[i][m_idx] not in ('', None, 0.0)]
                    avg = float(calculate_average(vals)) if vals else 0.0
                    avg_row.append(avg)
                start_row = row
                for idx in range(5):
                    if idx == 0:
                        ws.cell(row=row, column=1, value=display_name)
                    else:
                        ws.cell(row=row, column=1, value=None)
                    ws.cell(row=row, column=2, value=str(idx+1))
                    for m in range(5):
                        c = ws.cell(row=row, column=3+m, value=metric_rows[idx][m])
                        c.number_format = '0.00'
                    for col in range(1, 8):
                        ws.cell(row=row, column=col).border = thin_border
                        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    row += 1
                # Average row (bold, no topology name, merged with topology group)
                ws.cell(row=row, column=1, value=None)
                ws.cell(row=row, column=2, value='Average')
                for m in range(5):
                    c = ws.cell(row=row, column=3+m, value=avg_row[m])
                    c.number_format = '0.00'
                for col in range(2, 8):
                    ws.cell(row=row, column=col).border = thin_border
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                for col in range(1, 8):
                    ws.cell(row=row, column=col).font = Font(bold=True)
                ws.merge_cells(start_row=start_row, start_column=1, end_row=row, end_column=1)
                ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = light_gray_fill
                row += 1
            # Graphs: place after the table
            row += 2
            # Arrange graphs: 2 per row horizontally, then next row
            graphs_per_row = 2
            col_start = 2  # Start from column B
            for i, fig in enumerate(self.graph_figures):
                with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmpfile:
                    fig.savefig(tmpfile.name, bbox_inches='tight', dpi=100)
                    img = XLImage(tmpfile.name)
                    col_offset = (i % graphs_per_row)
                    graph_col = col_start + col_offset * 10  # Shift second graph a bit left
                    graph_cell = f'{get_column_letter(graph_col)}{row}'
                    ws.add_image(img, graph_cell)
                    if (i + 1) % graphs_per_row == 0:
                        row += 20  # Move to next row after 2 graphs
                    try:
                        os.remove(tmpfile.name)
                    except Exception:
                        pass
            if len(self.graph_figures) % graphs_per_row != 0:
                row += 20  # Add space for the last row if not full
            autofit_columns(ws)
            wb.save(filename)
            messagebox.showinfo("Success", f"Results saved to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save results to Excel: {str(e)}")

    def show_graphs(self):
        if not MATPLOTLIB_AVAILABLE:
            messagebox.showerror("Error", "matplotlib is not installed. Please install it using: pip install matplotlib")
            return
        self.generate_graph_figures()
        self.clear_container()
        # Heading at the very top, centered
        title_label = tk.Label(self.main_container, text="Network Performance Graph Analysis",
                               font=("Arial", 20, "bold"), bg='#f0f0f0', anchor='center', justify='center')
        title_label.pack(pady=(20, 0))
        # Legend below the heading, centered
        legend_frame = tk.Frame(self.main_container, bg='#f0f0f0')
        legend_frame.pack(pady=(0, 10))
        colors = ['#e41a1c', '#377eb8', '#4daf4a', '#ff9800']
        display_names = ['DMVPN', 'S2S VPN', 'No VPN', 'S2S VPN & DMVPN']
        for i, (color, name) in enumerate(zip(colors, display_names)):
            dot = tk.Canvas(legend_frame, width=18, height=18, bg='#f0f0f0', highlightthickness=0)
            dot.create_oval(3, 3, 15, 15, fill=color, outline=color)
            dot.pack(side='left', padx=(10 if i > 0 else 0, 2))
            label = tk.Label(legend_frame, text=name, font=("Arial", 12, "bold"), bg='#f0f0f0')
            label.pack(side='left', padx=(0, 10))
        # Use a canvas with a scrollbar for the graph area, with left padding
        canvas_frame = tk.Frame(self.main_container, bg='#f0f0f0')
        canvas_frame.pack(expand=True, fill='both')
        left_pad = tk.Frame(canvas_frame, width=60, bg='#f0f0f0')  # Add left padding
        left_pad.pack(side='left', fill='y')
        canvas = tk.Canvas(canvas_frame, bg='#f0f0f0', highlightthickness=0)
        vscrollbar = tk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f0f0f0')
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=vscrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        vscrollbar.pack(side="right", fill="y")
        # Enable mouse wheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        scrollable_frame.grid_columnconfigure(0, weight=1)
        scrollable_frame.grid_columnconfigure(1, weight=1)
        # Place each graph in its own frame
        for i, fig in enumerate(self.graph_figures):
            graph_container = tk.Frame(scrollable_frame, bg='#f0f0f0')
            graph_container.grid(row=(i//2), column=i%2, padx=10, pady=10, sticky='nsew')
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            canvas_tk = FigureCanvasTkAgg(fig, master=graph_container)
            canvas_tk.draw()
            canvas_tk.get_tk_widget().pack(expand=True, fill='both')
            scrollable_frame.grid_rowconfigure((i//2), weight=1)
        canvas.update_idletasks()
        canvas.xview_moveto(0)
        # Navigation and export buttons (always visible below graphs)
        btn_frame = tk.Frame(self.main_container, bg='#f0f0f0')
        btn_frame.pack(fill='x', padx=20, pady=10, side='bottom')
        left_btn_frame = tk.Frame(btn_frame, bg='#f0f0f0')
        left_btn_frame.pack(side='left')
        select_new_btn = tk.Button(left_btn_frame, text="Select New Files", command=self.setup_file_selection_gui,
                                   font=("Arial", 7, "bold"), bg='#2196F3', fg='white', padx=14, pady=5)
        select_new_btn.pack(side='left', padx=2)
        back_btn = tk.Button(left_btn_frame, text="← Back", command=self.show_results,
                             font=("Arial", 7, "bold"), bg='#9E9E9E', fg='white', padx=14, pady=5)
        back_btn.pack(side='left', padx=2)
        close_btn = tk.Button(left_btn_frame, text="✕ Close", command=self.close_application,
                              font=("Arial", 7, "bold"), bg='#F44336', fg='white', padx=14, pady=5)
        close_btn.pack(side='left', padx=2)
        export_frame = tk.Frame(btn_frame, bg='#f0f0f0')
        export_frame.pack(side='right')
        pdf_btn = tk.Button(export_frame, text="Save Results to PDF", command=self.save_results_to_pdf,
                            font=("Arial", 8), bg='#FF5722', fg='white', padx=14, pady=6)
        pdf_btn.pack(side='left', padx=5)
        excel_btn = tk.Button(export_frame, text="Save Results to Excel", command=self.save_results_to_excel,
                              font=("Arial", 8), bg='#9C27B0', fg='white', padx=14, pady=6)
        excel_btn.pack(side='left', padx=5)

    def run(self):
        self.root.mainloop()

# Main execution
if __name__ == "__main__":
    app = NetworkPerformanceExtractor()
    app.run() 